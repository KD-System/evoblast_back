"""
Сервис для работы с файлами.

Файлы загружаются напрямую в фиксированный индекс (SEARCH_INDEX_ID).
Индекс НЕ пересоздаётся при каждом изменении.
"""
import io
import logging
from typing import Dict, Any, List, Tuple, Optional
from fastapi import UploadFile

from app.database import mongodb
from app.services import yandex_service

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {'txt', 'pdf', 'doc', 'docx', 'md', 'json', 'csv', 'xls', 'xlsx'}
MAX_FILE_SIZE = 30 * 1024 * 1024  # 30 MB
MAX_FILES_PER_UPLOAD = 10


def extract_text_from_file(content: bytes, file_type: str) -> str:
    """Извлечь весь текст из файла"""
    try:
        # Текстовые файлы
        if file_type in ('txt', 'md', 'json', 'csv'):
            return content.decode('utf-8', errors='ignore')

        # PDF
        if file_type == 'pdf':
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(content))
                text_parts = []
                for page in reader.pages:
                    text_parts.append(page.extract_text() or "")
                return "\n".join(text_parts)
            except Exception as e:
                logger.warning(f"PDF extraction failed: {e}")
                return ""

        # DOCX
        if file_type == 'docx':
            try:
                from docx import Document
                doc = Document(io.BytesIO(content))
                text_parts = [p.text for p in doc.paragraphs]
                return "\n".join(text_parts)
            except Exception as e:
                logger.warning(f"DOCX extraction failed: {e}")
                return ""

        # XLSX
        if file_type == 'xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content), read_only=True)
                text_parts = []
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " | ".join(str(c) for c in row if c)
                        if row_text:
                            text_parts.append(row_text)
                return "\n".join(text_parts)
            except Exception as e:
                logger.warning(f"XLSX extraction failed: {e}")
                return ""

        # DOC, XLS — старые форматы, сложно извлечь
        if file_type in ('doc', 'xls'):
            return f"[Файл формата .{file_type} — превью недоступно]"

        return ""
    except Exception as e:
        logger.error(f"Text extraction error: {e}")
        return ""


def get_file_extension(filename: str) -> str:
    if '.' in filename:
        return filename.rsplit('.', 1)[1].lower()
    return ''


def is_allowed_file(filename: str) -> bool:
    return get_file_extension(filename) in ALLOWED_EXTENSIONS


async def upload_files(
    user_id: str,
    files: List[UploadFile],
    metadata: Dict[str, Any] = None
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Загрузить файлы в Yandex Cloud и добавить в индекс.
    Файлы добавляются в существующий индекс (SEARCH_INDEX_ID).
    """
    if len(files) > MAX_FILES_PER_UPLOAD:
        raise ValueError(f"Максимум {MAX_FILES_PER_UPLOAD} файлов за раз")

    uploaded_files = []
    errors = []

    for file in files:
        try:
            if not is_allowed_file(file.filename):
                errors.append(f"{file.filename}: неподдерживаемый тип")
                continue

            content = await file.read()
            file_size = len(content)

            if file_size > MAX_FILE_SIZE:
                errors.append(f"{file.filename}: слишком большой (макс. 30MB)")
                continue

            file_type = get_file_extension(file.filename)

            # Извлекаем текст для превью
            text_content = extract_text_from_file(content, file_type)

            # Загружаем в Yandex Cloud и добавляем в индекс
            yandex_file_id = await yandex_service.upload_file_to_index(content, file.filename)

            # Сохраняем запись в MongoDB со статусом "ready"
            file_record = await mongodb.create_file_record(
                user_id=user_id,
                filename=file.filename,
                file_type=file_type,
                file_size=file_size,
                yandex_file_id=yandex_file_id,
                content=text_content,
                binary_content="",
                metadata=metadata or {},
                status="ready"
            )

            # Сохраняем бинарный контент в GridFS
            await mongodb.gridfs_upload(file_record["file_id"], file.filename, content)

            uploaded_files.append(file_record)
            logger.info(f"✅ File uploaded and indexed: {file.filename} -> {yandex_file_id}")

        except Exception as e:
            logger.error(f"❌ Error uploading {file.filename}: {e}")
            errors.append(f"{file.filename}: {str(e)}")

    return uploaded_files, errors


async def get_all_files() -> List[Dict[str, Any]]:
    """Получить список ВСЕХ файлов (для всех пользователей)"""
    files = await mongodb.get_all_active_files()

    for f in files:
        if "_id" in f:
            del f["_id"]
        if "content" in f:
            del f["content"]
        if "binary_content" in f:
            del f["binary_content"]

    return files


async def get_user_files(user_id: str) -> List[Dict[str, Any]]:
    """Получить список файлов конкретного пользователя"""
    files = await mongodb.get_user_files(user_id)

    for f in files:
        if "_id" in f:
            del f["_id"]
        if "content" in f:
            del f["content"]
        if "binary_content" in f:
            del f["binary_content"]

    return files


async def get_file(file_id: str) -> Dict[str, Any]:
    """Получить файл по ID"""
    file = await mongodb.get_file_by_id(file_id)

    if not file:
        raise ValueError(f"Файл не найден: {file_id}")

    if "_id" in file:
        del file["_id"]

    return file


async def delete_file(file_id: str) -> bool:
    """Удалить файл из индекса и базы данных"""
    file = await mongodb.delete_file_record(file_id)

    if not file:
        raise ValueError(f"Файл не найден: {file_id}")

    # Удаляем из Yandex Cloud (индекс + storage)
    if file.get("yandex_file_id"):
        await yandex_service.delete_file_from_index(file["yandex_file_id"])

    # Удаляем из GridFS
    await mongodb.gridfs_delete(file_id)

    logger.info(f"🗑️ File deleted: {file_id}")
    return True


async def delete_all_files() -> int:
    """Удалить ВСЕ файлы из индекса и базы данных"""
    files = await mongodb.get_all_active_files()

    # Удаляем каждый файл из Yandex Cloud и GridFS
    for file in files:
        if file.get("yandex_file_id"):
            try:
                await yandex_service.delete_file_from_index(file["yandex_file_id"])
            except Exception as e:
                logger.warning(f"⚠️ Failed to delete file from index: {e}")

        try:
            await mongodb.gridfs_delete(file["file_id"])
        except Exception:
            pass

    # Помечаем все как удалённые в MongoDB
    deleted_count = await mongodb.delete_all_files()

    logger.info(f"🗑️ Deleted all {deleted_count} files")
    return deleted_count


async def get_index_info() -> Dict[str, Any]:
    """Получить информацию об индексе"""
    return await yandex_service.get_index_info()


async def list_index_files(limit: int = 100) -> List[Dict[str, Any]]:
    """Получить список файлов в индексе"""
    return await yandex_service.list_index_files(limit)
